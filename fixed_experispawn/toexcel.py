#!/usr/bin/python3

import sys
import time
import pandas
from openpyxl import load_workbook

# B		ARCH		LANG		#TH		#RUNS	FCN		AVERAGE

# 01C		gal		cilk		32		100		slow		0.000000001
# 02C

def panda_to_excel(bmk, arch, lang, nth, nruns, fcn, avg, timer=''):

    # dictionary of data  
    dct = {'BNCHMRK': [bmk],  
        'TIMER': [timer], # optional
        'ARCH': [arch], 
        'LANG':[lang],
        '#TH':[nth],
        '#RUNS':[nruns],
        'FCN':[fcn],
        'AVG':[avg],
    }
    #dct = [[bmk, arch, lang, nth, nruns, fun, avg, timer]] 

    # 

    ss = '../fixedexperispawn_data.xlsx'
    sheetname="Sheet4"

    # forming dataframe 
    df = pandas.DataFrame(data=dct)  

    # determine the starting row using openpyxl
    # This prevents overwriting existing data inside the sheet
    wb = load_workbook(ss)
    startrw = wb[sheetname].max_row if sheetname in wb.sheetnames else 1

    # open ExcelWriter in append mode and overlay the data
    with pandas.ExcelWriter(ss, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel( writer, sheet_name=sheetname, startrow=startrw, index=False, header=(startrw == 1))

    # # Read existing data
    # reader = pandas.read_excel(ss, sheet_name=['Sheet1'])

    # print(type(list(reader.values())[0]))
    # print(len(reader),list(reader.values())[0].shape[0])

    # writer = pandas.ExcelWriter(ss, mode='a',if_sheet_exists='overlay')
    # df.to_excel(writer, sheet_name='Sheet1', index=False, header=False,startrow=list(reader.values())[0].shape[0]+1)
    # writer.close()

    # df.to_excel(file_name)
    print('DataFrame is written to Excel File successfully.')

    return


def main():

    panda_to_excel(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4],sys.argv[5],sys.argv[6],sys.argv[7],sys.argv[8])

    return

if __name__ == '__main__':
    main()
